#!/usr/bin/env python3
"""Create an Excel workbook with charts from processed CSVs."""
from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.utils import get_column_letter

BASE = Path('/Users/benledwon/Desktop/Github_connection/bike_share')
PROCESSED = BASE / 'data' / 'processed'
OUT = BASE / 'figures' / 'cyclistic_2019_charts.xlsx'

wb = Workbook()
wb.remove(wb.active)


def _to_number(value: str):
    if value is None:
        return value
    v = str(value).strip()
    if v == '':
        return v
    # Try int then float
    try:
        if '.' not in v and 'e' not in v.lower():
            return int(v)
    except ValueError:
        pass
    try:
        return float(v)
    except ValueError:
        return value


def add_sheet_from_csv(name: str, csv_path: Path, drop_unknown=False):
    ws = wb.create_sheet(title=name)
    with csv_path.open() as f:
        reader = csv.reader(f)
        for r_idx, row in enumerate(reader):
            if r_idx == 0:
                ws.append(row)
                continue
            if drop_unknown and len(row) > 1 and str(row[1]).strip().lower() == 'unknown':
                continue
            ws.append([_to_number(c) for c in row])
    return ws

# Load data sheets
ws_overall = add_sheet_from_csv('summary_overall', PROCESSED / 'summary_overall.csv', drop_unknown=True)
ws_day = add_sheet_from_csv('rides_by_day_user', PROCESSED / 'rides_by_day_user.csv', drop_unknown=True)
ws_avg_day = add_sheet_from_csv('avg_ride_by_day', PROCESSED / 'avg_ride_seconds_by_day_user.csv', drop_unknown=True)
ws_hour = add_sheet_from_csv('rides_by_hour', PROCESSED / 'rides_by_hour_user.csv', drop_unknown=True)
ws_month = add_sheet_from_csv('rides_by_month', PROCESSED / 'rides_by_month_user.csv', drop_unknown=True)
ws_commute = add_sheet_from_csv('commute_share', PROCESSED / 'commute_share_weekday.csv', drop_unknown=True)
ws_round = add_sheet_from_csv('round_trip_share', PROCESSED / 'round_trip_share.csv', drop_unknown=True)
ws_long = add_sheet_from_csv('long_ride_share', PROCESSED / 'long_ride_share.csv', drop_unknown=True)

# Create a Charts sheet
ws_charts = wb.create_sheet(title='Charts')

# Helper: build a pivot-like table for charts

DAY_FULL = {
    'Mon': 'Monday',
    'Tue': 'Tuesday',
    'Wed': 'Wednesday',
    'Thu': 'Thursday',
    'Fri': 'Friday',
    'Sat': 'Saturday',
    'Sun': 'Sunday',
}


def build_pivot(ws_src, key_col, series_col, value_col, target_row):
    data = {}
    keys = []
    series = []
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        k = row[key_col]
        if isinstance(k, str) and k in DAY_FULL:
            k = DAY_FULL[k]
        s = row[series_col]
        if isinstance(s, str):
            s = s.capitalize()
        v = row[value_col]
        data.setdefault(k, {})[s] = v
        if k not in keys:
            keys.append(k)
        if s not in series:
            series.append(s)
    # header
    ws_charts.cell(row=target_row, column=1, value='key')
    for idx, s in enumerate(series, start=2):
        ws_charts.cell(row=target_row, column=idx, value=s)
    # rows
    for r, k in enumerate(keys, start=target_row+1):
        ws_charts.cell(row=r, column=1, value=k)
        for c, s in enumerate(series, start=2):
            ws_charts.cell(row=r, column=c, value=data.get(k, {}).get(s, 0))
    return target_row, target_row+len(keys), 1, 1+len(series)

row = 1

def style_chart(chart, width=26, height=14):
    chart.width = width
    chart.height = height
    chart.style = 10
    chart.legend.position = 'r'
    chart.legend.overlay = False
    chart.y_axis.majorGridlines = ChartLines()
    chart.x_axis.tickLblPos = 'low'
    chart.y_axis.tickLblPos = 'low'
    return chart

# Chart 1: Rides by Day of Week (counts)
start, end, c1, c2 = build_pivot(ws_day, 0, 1, 2, row)
chart1 = BarChart()
chart1.title = 'Rides by Day of Week'
chart1.y_axis.title = 'Rides'
chart1.x_axis.title = 'Day'
chart1.grouping = 'clustered'
chart1.gapWidth = 120
chart1.y_axis.number_format = '0'
chart1.add_data(Reference(ws_charts, min_col=2, min_row=start, max_row=end, max_col=c2), titles_from_data=True)
chart1.set_categories(Reference(ws_charts, min_col=1, min_row=start+1, max_row=end))
style_chart(chart1)
ws_charts.add_chart(chart1, f'A{end+2}')
row = end + 20

# Chart 2: Avg Ride Length by Day of Week
start, end, c1, c2 = build_pivot(ws_avg_day, 0, 1, 2, row)
chart2 = BarChart()
chart2.title = 'Avg Ride Length (sec) by Day'
chart2.y_axis.title = 'Seconds'
chart2.x_axis.title = 'Day'
chart2.grouping = 'clustered'
chart2.gapWidth = 120
chart2.y_axis.number_format = '0'
chart2.add_data(Reference(ws_charts, min_col=2, min_row=start, max_row=end, max_col=c2), titles_from_data=True)
chart2.set_categories(Reference(ws_charts, min_col=1, min_row=start+1, max_row=end))
style_chart(chart2)
ws_charts.add_chart(chart2, f'A{end+2}')
row = end + 20

# Chart 3: Rides by Month
start, end, c1, c2 = build_pivot(ws_month, 0, 1, 2, row)
chart3 = LineChart()
chart3.title = 'Rides by Month'
chart3.y_axis.title = 'Rides'
chart3.x_axis.title = 'Month'
chart3.smooth = True
chart3.y_axis.number_format = '0'
chart3.add_data(Reference(ws_charts, min_col=2, min_row=start, max_row=end, max_col=c2), titles_from_data=True)
chart3.set_categories(Reference(ws_charts, min_col=1, min_row=start+1, max_row=end))
style_chart(chart3)
ws_charts.add_chart(chart3, f'A{end+2}')
row = end + 20

# Chart 4: Rides by Hour
start, end, c1, c2 = build_pivot(ws_hour, 0, 1, 2, row)
chart4 = LineChart()
chart4.title = 'Rides by Hour'
chart4.y_axis.title = 'Rides'
chart4.x_axis.title = 'Hour'
chart4.smooth = True
chart4.y_axis.number_format = '0'
chart4.add_data(Reference(ws_charts, min_col=2, min_row=start, max_row=end, max_col=c2), titles_from_data=True)
chart4.set_categories(Reference(ws_charts, min_col=1, min_row=start+1, max_row=end))
style_chart(chart4)
ws_charts.add_chart(chart4, f'A{end+2}')
row = end + 20

# Chart 5: Weekend vs Weekday share
# Build a 2-row table for share
ws_charts.cell(row=row, column=1, value='user_type')
ws_charts.cell(row=row, column=2, value='weekend_share')
ws_charts.cell(row=row, column=3, value='weekday_share')

# Read from summary_overall
summary = {}
for r in ws_overall.iter_rows(min_row=2, values_only=True):
    summary[r[0]] = r

for i, user in enumerate(['member','casual'], start=row+1):
    weekend = float(summary[user][5])
    weekday = float(summary[user][6])
    total = weekend + weekday
    ws_charts.cell(row=i, column=1, value=user)
    ws_charts.cell(row=i, column=2, value=weekend/total if total else 0)
    ws_charts.cell(row=i, column=3, value=weekday/total if total else 0)

chart5 = BarChart()
chart5.title = 'Weekend vs Weekday Share'
chart5.y_axis.title = 'Share'
chart5.x_axis.title = 'User Type'
chart5.add_data(Reference(ws_charts, min_col=2, min_row=row, max_row=row+2, max_col=3), titles_from_data=True)
chart5.set_categories(Reference(ws_charts, min_col=1, min_row=row+1, max_row=row+2))
chart5.type = 'col'
chart5.grouping = 'stacked'
chart5.y_axis.number_format = '0%'
chart5.y_axis.scaling.min = 0
chart5.y_axis.scaling.max = 1
style_chart(chart5, width=22, height=12)
ws_charts.add_chart(chart5, f'A{row+4}')
row = row + 20

# Chart 6: Long ride share (>30m, >60m)
ws_charts.cell(row=row, column=1, value='user_type')
ws_charts.cell(row=row, column=2, value='share_over_30m')
ws_charts.cell(row=row, column=3, value='share_over_60m')

long_share = {}
for r in ws_long.iter_rows(min_row=2, values_only=True):
    long_share[r[0]] = r

for i, user in enumerate(['member','casual'], start=row+1):
    ws_charts.cell(row=i, column=1, value=user)
    ws_charts.cell(row=i, column=2, value=float(long_share[user][4]))
    ws_charts.cell(row=i, column=3, value=float(long_share[user][5]))

chart6 = BarChart()
chart6.title = 'Long Ride Share'
chart6.y_axis.title = 'Share'
chart6.x_axis.title = 'User Type'
chart6.add_data(Reference(ws_charts, min_col=2, min_row=row, max_row=row+2, max_col=3), titles_from_data=True)
chart6.set_categories(Reference(ws_charts, min_col=1, min_row=row+1, max_row=row+2))
chart6.y_axis.number_format = '0%'
chart6.y_axis.scaling.min = 0
chart6.y_axis.scaling.max = 1
style_chart(chart6, width=22, height=12)
ws_charts.add_chart(chart6, f'A{row+4}')

# Save
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f'Wrote {OUT}')
